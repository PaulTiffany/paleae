
# paleae_one.py — single-file snapshot/feed tool
# Version: 1.1.0-singlefile
# Generated in-session. No external dependencies required; optional: tiktoken, openpyxl.
# Usage (examples):
#   python paleae_one.py snapshot . --format xlsx --out snapgrid.xlsx
#   python paleae_one.py feed . --format jsonl --out feed.jsonl --diff HEAD~1
#   python paleae_one.py snapshot . --format json --diff HEAD~1 --diff-summary
#   python paleae_one.py check . --profile ai_optimized --out paleae_check.json

import argparse
import ast
import csv
import hashlib
import json
import os
import re
import subprocess
import sys
import time
import zipfile
from dataclasses import asdict, dataclass
from typing import Any, Dict, List, Optional, Tuple

paleae_VERSION = "1.1.0-singlefile"

try:
    import tiktoken  # type: ignore
    _HAS_TIKTOKEN = True
except Exception:
    _HAS_TIKTOKEN = False

try:
    from openpyxl import Workbook  # noqa: F401
    _HAS_OPENPYXL = True
except Exception:
    _HAS_OPENPYXL = False

def sha256_text(s: str) -> str:
    return hashlib.sha256(s.encode("utf-8", errors="ignore")).hexdigest()

def estimate_tokens(s: str) -> int:
    if _HAS_TIKTOKEN:
        try:
            enc = tiktoken.get_encoding("cl100k_base")
            return len(enc.encode(s))
        except Exception:
            pass
    # crude fallback: ~4 chars/token
    return max(1, len(s) // 4)

def which_git() -> Optional[str]:
    from shutil import which
    return which("git")

def is_text_file(path: str) -> bool:
    # Heuristic: we consider known text extensions as text; otherwise try a small read for binary bytes
    text_exts = {
        ".py",".md",".rst",".txt",".toml",".ini",".cfg",".json",".yml",".yaml",".xml",".csv",".tsv",".html",".css",".js",".ts",".tsx",".pyi"
    }
    ext = os.path.splitext(path)[1].lower()
    if ext in text_exts:
        return True
    try:
        with open(path, "rb") as f:
            chunk = f.read(1024)
        if b"\\x00" in chunk:
            return False
        # Assume text if decodes as utf-8
        chunk.decode("utf-8")
        return True
    except Exception:
        return False

def read_text(path: str) -> str:
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        return f.read()

def language_from_path(path: str) -> str:
    ext = os.path.splitext(path)[1].lower()
    return {
        ".py": "python",
        ".md": "markdown",
        ".rst": "rst",
        ".json": "json",
        ".toml": "toml",
        ".yaml": "yaml",
        ".yml": "yaml",
        ".ini": "ini",
        ".cfg": "ini",
        ".xml": "xml",
        ".csv": "csv",
        ".tsv": "tsv",
        ".js": "javascript",
        ".ts": "typescript",
        ".tsx": "typescriptreact",
        ".html": "html",
        ".css": "css",
        ".pyi": "python-stub",
    }.get(ext, "text")

def kind_from_path(path: str) -> str:
    low = path.replace("\\\\", "/").lower()
    if low.startswith("tests/") or "/tests/" in low or low.endswith("_test.py") or low.startswith("test_") or "/test_" in low:
        return "test"
    ext = os.path.splitext(low)[1]
    if ext in {".py",".pyi",".js",".ts",".tsx",".css",".html"}:
        return "code"
    if ext in {".md",".rst",".txt"}:
        return "doc"
    if ext in {".json",".toml",".yml",".yaml",".ini",".cfg",".xml"}:
        return "config"
    return "data"

DEFAULT_EXCLUDES = [
    r"(^|/)\\.git($|/)", r"(^|/)\\.hg($|/)", r"(^|/)\\.svn($|/)",
    r"(^|/)__pycache__($|/)",
    r"(^|/)\\.mypy_cache($|/)", r"(^|/)\\.pytest_cache($|/)",
    r"(^|/)\\.venv($|/)", r"(^|/)venv($|/)", r"(^|/)env($|/)",
    r"(^|/)artifacts($|/)", r"(^|/)site($|/)", r"(^|/)build($|/)",
    r"(^|/)coverage($|/)", r"(^|/)htmlcov($|/)",
    r"(^|/)\\.coverage", r"(^|/)\\.env($|/)",
    r"(^|/)dist($|/)",
]

PROFILE_INCLUDES = {
    "minimal": [r".*"],
    "ai_optimized": [
        r"^(pylantern|tests|pyproject\\.toml|README(\\.md|\\.rst)?|ROADMAP\\.md|CHANGELOG\\.md|CONTRIBUTING\\.md)(/.*)?$",
        r"^reports/(pytest_results\\.xml|coverage\\.xml)$",
    ],
}

PROFILE_EXCLUDES = {
    "minimal": DEFAULT_EXCLUDES,
    "ai_optimized": DEFAULT_EXCLUDES + [
        r"(^|/)docs/",
        r"(^|/)artifacts/",
        r"(^|/)site/",
    ],
}

def compile_patterns(patterns: List[str]) -> List[re.Pattern]:
    return [re.compile(p) for p in patterns]

def path_matches_any(path: str, patterns: List[re.Pattern]) -> bool:
    return any(p.search(path) for p in patterns)

def py_deps_from_source(source: str) -> List[str]:
    try:
        tree = ast.parse(source)
    except Exception:
        return []
    deps: List[str] = []
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            for alias in node.names:
                deps.append(alias.name.split(".")[0])
        elif isinstance(node, ast.ImportFrom):
            if node.module:
                deps.append(node.module.split(".")[0])
    # Deduplicate, stable order
    out = []
    seen = set()
    for d in deps:
        if d not in seen:
            out.append(d)
            seen.add(d)
    return out

def git_changed_files(root: str, ref: str) -> Optional[List[str]]:
    exe = which_git()
    if not exe:
        return None
    try:
        # returns paths relative to root
        res = subprocess.run(
            [exe, "diff", "--name-only", ref, "HEAD"],
            cwd=root, capture_output=True, text=True, check=True
        )
        files = [line.strip() for line in res.stdout.splitlines() if line.strip()]
        return files
    except Exception:
        return None

def chunk_text_by_lines(text: str, max_chars: int) -> List[Tuple[int,int,str]]:
    lines = text.splitlines()
    chunks: List[Tuple[int,int,str]] = []
    buf: List[str] = []
    start = 1
    cur_len = 0
    for i, line in enumerate(lines, start=1):
        add = len(line) + 1  # account for newline
        if cur_len + add > max_chars and buf:
            chunk_text = "\\n".join(buf)
            chunks.append((start, i-1, chunk_text))
            buf = [line]
            start = i
            cur_len = len(line) + 1
        else:
            buf.append(line)
            cur_len += add
    if buf:
        chunks.append((start, len(lines), "\\n".join(buf)))
    if not chunks:
        chunks.append((1, 1, text))
    return chunks

def chunk_python_by_defs(text: str, max_chars: int) -> List[Tuple[int,int,str]]:
    # Attempt to split by top-level class/def; fallback to lines
    try:
        tree = ast.parse(text)
    except Exception:
        return chunk_text_by_lines(text, max_chars)
    boundaries = []
    for node in tree.body:
        if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef, ast.ClassDef)):
            lineno = getattr(node, "lineno", None)
            end_lineno = getattr(node, "end_lineno", None)
            if lineno is not None and end_lineno is not None:
                boundaries.append((lineno, end_lineno))
    if not boundaries:
        return chunk_text_by_lines(text, max_chars)
    boundaries.sort()
    chunks: List[Tuple[int,int,str]] = []
    for (s,e) in boundaries:
        # slice lines to enforce max_chars
        segment = "\\n".join(text.splitlines()[s-1:e])
        if len(segment) <= max_chars:
            chunks.append((s,e,segment))
        else:
            # further split by lines
            for (ss, ee, seg) in chunk_text_by_lines(segment, max_chars):
                chunks.append((s+(ss-1), s+(ee-1), seg))
    return chunks

def count_words(s: str) -> int:
    return len(s.split()) if s else 0

SNAPGRID_COLUMNS = [
    "id","relpath","kind","language","start","end","bytes","line_count","word_count","est_tokens","sha256","deps","tags","profile","git_ref","text"
]

@dataclass
class Row:
    id: str
    relpath: str
    kind: str
    language: str
    start: int
    end: int
    bytes: int
    line_count: int
    word_count: int
    est_tokens: int
    sha256: str
    deps: str
    tags: str
    profile: str
    git_ref: str
    text: str

def row_from_chunk(relpath: str, profile: str, git_ref: str, kind: str, language: str, start: int, end: int, text: str, deps_list: list[str]) -> Row:
    deps_str = ";".join(deps_list) if deps_list else ""
    _id = f"{relpath}:{start}-{end}"
    return Row(
        id=_id,
        relpath=relpath,
        kind=kind,
        language=language,
        start=start,
        end=end,
        bytes=len(text.encode("utf-8", errors="ignore")),
        line_count=text.count("\\n") + 1 if text else 0,
        word_count=count_words(text),
        est_tokens=estimate_tokens(text),
        sha256=sha256_text(text),
        deps=deps_str,
        tags="",
        profile=profile,
        git_ref=git_ref,
        text=text
    )

def write_json(path: str, data: Dict[str, Any]) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

def write_markdown(path: str, rows: List[Row], meta: Dict[str, Any]) -> None:
    with open(path, "w", encoding="utf-8") as f:
        f.write("# paleae Snapshot (v" + paleae_VERSION + ")\\n\\n")
        f.write("## Run Meta\\n\\n")
        for k,v in meta.items():
            f.write(f"- **{k}**: {v}\\n")
        f.write("\\n---\\n\\n")
        for r in rows:
            f.write(f"## `{r.relpath}` [{r.start}-{r.end}] — {r.kind}/{r.language}\\n\\n")
            f.write(f"- id: `{r.id}`  \\n- bytes: {r.bytes}  \\n- lines: {r.line_count}  \\n- words: {r.word_count}  \\n- est_tokens: {r.est_tokens}  \\n- sha256: `{r.sha256}`  \\n")
            if r.deps:
                f.write(f"- deps: `{r.deps}`  \\n")
            f.write("\\n```text\\n")
            f.write(r.text)
            f.write("\\n```\\n\\n")

def write_csv(path: str, rows: List[Row]) -> None:
    with open(path, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f, quoting=csv.QUOTE_ALL)
        w.writerow(SNAPGRID_COLUMNS)
        for r in rows:
            dd = asdict(r)
            w.writerow([dd[c] for c in SNAPGRID_COLUMNS])

def write_xlsx(path: str, rows: List[Row], meta: Dict[str, Any]) -> None:
    if not _HAS_OPENPYXL:
        raise RuntimeError("openpyxl not installed; cannot write .xlsx")
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = "snapgrid"
    ws.append(SNAPGRID_COLUMNS)
    for r in rows:
        dd = asdict(r)
        ws.append([dd[c] for c in SNAPGRID_COLUMNS])
    # reasonable column widths
    for idx, col in enumerate(SNAPGRID_COLUMNS, start=1):
        width = 12
        if col in ("relpath", "deps", "text"):
            width = 80 if col == "text" else 40
        ws.column_dimensions[get_column_letter(idx)].width = width
    # meta sheet
    meta_ws = wb.create_sheet("run_meta")
    for k,v in meta.items():
        meta_ws.append([k, json.dumps(v) if isinstance(v,(dict,list)) else v])
    wb.save(path)

def write_jsonl(path: str, rows: List[Row]) -> None:
    with open(path, "w", encoding="utf-8") as f:
        for r in rows:
            dd = asdict(r)
            # jsonl excludes meta-ish tags for machine use? Keep all except 'text' depending on need.
            f.write(json.dumps(dd, ensure_ascii=False) + "\\n")

def make_zip(zip_path: str, files: Dict[str, bytes]) -> None:
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as z:
        for arcname, data in files.items():
            z.writestr(arcname, data)

def collect_files(root: str, include: List[re.Pattern], exclude: List[re.Pattern], only_changed: Optional[List[str]] = None) -> List[str]:
    root = os.path.abspath(root)
    rels: List[str] = []
    changed_set = set(os.path.normpath(p) for p in (only_changed or []))
    for dirpath, dirnames, filenames in os.walk(root):
        # normalize path with forward slashes for regex
        rel_dir = os.path.relpath(dirpath, root).replace("\\\\", "/")
        if rel_dir == ".":
            rel_dir = ""
        # skip excluded dirs quickly
        skip_dir = False
        for p in exclude:
            if p.search(rel_dir + "/"):
                skip_dir = True
                break
        if skip_dir:
            dirnames[:] = []  # don't descend
            continue
        for fn in sorted(filenames):
            rel = os.path.join(rel_dir, fn).replace("\\\\", "/") if rel_dir else fn
            full = os.path.join(root, rel)
            if only_changed is not None and os.path.normpath(rel) not in changed_set:
                continue
            if path_matches_any(rel, exclude):
                continue
            if not path_matches_any(rel, include):
                continue
            if not os.path.isfile(full):
                continue
            if not is_text_file(full):
                continue
            rels.append(rel)
    rels.sort()
    return rels

def generate_rows(root: str, files: List[str], profile: str, chunk_by: str, max_chars: int, git_ref: str) -> List[Row]:
    rows: List[Row] = []
    for rel in files:
        full = os.path.join(root, rel)
        try:
            text = read_text(full)
        except Exception:
            continue
        language = language_from_path(rel)
        kind = kind_from_path(rel)
        deps_list: List[str] = []
        if language == "python":
            deps_list = py_deps_from_source(text)
        # chunk
        chunks: List[Tuple[int,int,str]]
        if language == "python" and chunk_by in ("functions","classes"):
            # classes == functions for this simple cut (both top-level defs)
            chunks = chunk_python_by_defs(text, max_chars)
        elif chunk_by == "lines":
            chunks = chunk_text_by_lines(text, max_chars)
        else:
            # file: single chunk (split if too big by lines anyway)
            if len(text) > max_chars:
                chunks = chunk_text_by_lines(text, max_chars)
            else:
                end_line = len(text.splitlines())
                chunks = [(1, end_line if end_line>0 else 1, text)]
        for (s,e,chunk_text) in chunks:
            rows.append(row_from_chunk(rel, profile, git_ref, kind, language, s, e, chunk_text, deps_list))
    # stable sort by relpath then start
    rows.sort(key=lambda r: (r.relpath, r.start, r.end))
    return rows

def diff_summary_payload(root: str, files: List[str]) -> Dict[str, Any]:
    out: List[Dict[str, Any]] = []
    for rel in files:
        full = os.path.join(root, rel)
        try:
            text = read_text(full)
        except Exception:
            continue
        out.append({"relpath": rel, "sha256": sha256_text(text)})
    return {"ts": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
            "count": len(out),
            "entries": out}

def run_snapshot(args: argparse.Namespace) -> int:
    root = args.root
    profile = args.profile
    includes = compile_patterns(PROFILE_INCLUDES.get(profile, PROFILE_INCLUDES["minimal"]))
    excludes = compile_patterns(PROFILE_EXCLUDES.get(profile, PROFILE_EXCLUDES["minimal"]))
    gitref = "HEAD"
    changed = None
    if args.diff:
        ch = git_changed_files(root, args.diff)
        if ch is None:
            print("[paleae] git not available or diff failed; proceeding without diff filter.", file=sys.stderr)
        else:
            changed = ch
            gitref = f"{args.diff}..HEAD"
    files = collect_files(root, includes, excludes, changed)
    rows = generate_rows(root, files, profile, args.chunk_by, args.max_chars, gitref)

    meta = {
        "tool": "paleae",
        "version": paleae_VERSION,
        "root": os.path.abspath(root),
        "profile": profile,
        "args": {k: v for k, v in vars(args).items() if k != 'func'},
        "ts": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        "counts": {
            "files": len(files),
            "rows": len(rows),
        }
    }

    out = args.out or ("snapshot." + args.format)
    fmt = args.format.lower()
    os.makedirs(os.path.dirname(out) or ".", exist_ok=True)

    if fmt == "json":
        data = {"meta": meta, "rows": [asdict(r) for r in rows]}
        write_json(out, data)
    elif fmt == "markdown":
        write_markdown(out, rows, meta)
    elif fmt == "xlsx":
        write_xlsx(out, rows, meta)
    elif fmt == "zip":
        # zip contains json + markdown
        json_data = {"meta": meta, "rows": [asdict(r) for r in rows]}
        json_bytes = json.dumps(json_data, indent=2, ensure_ascii=False).encode("utf-8")
        # produce markdown string into bytes
        md_content = ["# paleae Snapshot (zip)\\n\\n","## Run Meta\\n\\n"]
        for k,v in meta.items():
            md_content.append(f"- **{k}**: {v}\\n")
        md_content.append("\\n---\\n\\n")
        for r in rows:
            md_content.append(f"## `{r.relpath}` [{r.start}-{r.end}] — {r.kind}/{r.language}\\n\\n")
            md_content.append(f"- id: `{r.id}`  \\n- bytes: {r.bytes}  \\n- lines: {r.line_count}  \\n- words: {r.word_count}  \\n- est_tokens: {r.est_tokens}  \\n- sha256: `{r.sha256}`  \\n")
            if r.deps:
                md_content.append(f"- deps: `{r.deps}`  \\n")
            md_content.append("\\n```text\\n")
            md_content.append(r.text)
            md_content.append("\\n```\\n\\n")
        md_bytes = "".join(md_content).encode("utf-8")
        files_map = {"snapshot.json": json_bytes, "SNAPSHOT.md": md_bytes}
        make_zip(out, files_map)
    else:
        print(f"Unsupported format: {fmt}", file=sys.stderr)
        return 2

    if args.diff_summary:
        # sidecar JSON with sha256 per (changed) file
        target_files = files if changed is None else changed
        payload = diff_summary_payload(root, target_files)
        base = out
        if "." in base:
            base = ".".join(base.split(".")[:-1])
        sidecar = base + ".diff.json"
        write_json(sidecar, payload)
        print(f"[paleae] wrote diff summary to {sidecar} ({payload['count']} entries).", file=sys.stderr)

    print(f"[paleae] wrote {fmt} to {out} ({len(rows)} rows from {len(files)} files).", file=sys.stderr)
    return 0

def run_feed(args: argparse.Namespace) -> int:
    root = args.root
    profile = args.profile
    includes = compile_patterns(PROFILE_INCLUDES.get(profile, PROFILE_INCLUDES["minimal"]))
    excludes = compile_patterns(PROFILE_EXCLUDES.get(profile, PROFILE_EXCLUDES["minimal"]))
    gitref = "HEAD"
    changed = None
    if args.diff:
        ch = git_changed_files(root, args.diff)
        if ch is None:
            print("[paleae feed] git not available or diff failed; proceeding without diff filter.", file=sys.stderr)
        else:
            changed = ch
            gitref = f"{args.diff}..HEAD"
    files = collect_files(root, includes, excludes, changed)
    rows = generate_rows(root, files, profile, args.chunk_by, args.max_chars, gitref)

    # Deterministic JSONL/CSV/XLSX
    out = args.out or ("feed." + args.format)
    fmt = args.format.lower()
    os.makedirs(os.path.dirname(out) or ".", exist_ok=True)
    if fmt == "jsonl":
        write_jsonl(out, rows)
    elif fmt == "csv":
        write_csv(out, rows)
    elif fmt == "xlsx":
        write_xlsx(out, rows, {"tool":"paleae-feed","version":paleae_VERSION})
    else:
        print(f"Unsupported format: {fmt}", file=sys.stderr)
        return 2

    if args.diff_summary:
        target_files = files if changed is None else changed
        payload = diff_summary_payload(root, target_files)
        base = out
        if "." in base:
            base = ".".join(base.split(".")[:-1])
        sidecar = base + ".diff.json"
        write_json(sidecar, payload)
        print(f"[paleae feed] wrote diff summary to {sidecar} ({payload['count']} entries).", file=sys.stderr)

    print(f"[paleae feed] wrote {fmt} to {out} ({len(rows)} rows from {len(files)} files).", file=sys.stderr)
    return 0

def run_check(args: argparse.Namespace) -> int:
    root = args.root
    profile = args.profile
    includes = compile_patterns(PROFILE_INCLUDES.get(profile, PROFILE_INCLUDES["minimal"]))
    excludes = compile_patterns(PROFILE_EXCLUDES.get(profile, PROFILE_EXCLUDES["minimal"]))
    files = collect_files(root, includes, excludes, None)
    errors: List[Dict[str, Any]] = []
    ok = 0
    for rel in files:
        full = os.path.join(root, rel)
        try:
            _ = read_text(full)
            ok += 1
        except Exception as e:
            errors.append({"relpath": rel, "error": repr(e)})
    report = {
        "tool": "paleae",
        "version": paleae_VERSION,
        "ts": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        "profile": profile,
        "counts": {"scanned": len(files), "readable": ok, "errors": len(errors)},
        "errors": errors[:20],  # cap
    }
    if args.out:
        os.makedirs(os.path.dirname(args.out) or ".", exist_ok=True)
        write_json(args.out, report)
        print(f"[paleae check] wrote report to {args.out}", file=sys.stderr)
    else:
        print(json.dumps(report, indent=2), file=sys.stdout)
    return 0 if not errors else 1

def build_arg_parser() -> argparse.ArgumentParser:
    tagline = "paleae: snapshots and feeds for repos, in human-friendly and machine-friendly forms."
    p = argparse.ArgumentParser(
        prog="paleae",
        description="Single-file repository snapshots & feeds — compact, fast, harmless.",
        epilog=tagline)
    p.add_argument("--version", action="store_true", help="Print version and exit")
    sub = p.add_subparsers(dest="cmd", required=False)

    def add_common(sp: argparse.ArgumentParser):
        sp.add_argument("root", nargs="?", default=".", help="Root directory")
        sp.add_argument("--profile", default="minimal", choices=list(PROFILE_INCLUDES.keys()), help="Inclusion profile")
        sp.add_argument("--include", action="append", default=None, help="Additional include regex (repeatable)")
        sp.add_argument("--exclude", action="append", default=None, help="Additional exclude regex (repeatable)")
        sp.add_argument("--diff", default=None, help="Git ref to diff against (e.g., HEAD~1)")
        sp.add_argument("--diff-summary", action="store_true", help="Emit a sidecar JSON with relpath+sha256 for (changed) files")
        sp.add_argument("--chunk-by", default="lines", choices=["file","lines","functions","classes"], help="Chunking strategy")
        sp.add_argument("--max-chars", type=int, default=16000, help="Max chars per chunk (keeps XLSX cells safe)")
        sp.add_argument("--out", default=None, help="Output path")

    ps = sub.add_parser("snapshot", help="Human-friendly snapshot (JSON/Markdown/XLSX/ZIP)")
    add_common(ps)
    ps.add_argument("--format", default="json", choices=["json","markdown","xlsx","zip"], help="Output format")
    ps.set_defaults(func=run_snapshot)

    pf = sub.add_parser("feed", help="Deterministic machine feed (JSONL/CSV/XLSX)")
    add_common(pf)
    pf.add_argument("--format", default="jsonl", choices=["jsonl","csv","xlsx"], help="Output format")
    pf.set_defaults(func=run_feed)

    pc = sub.add_parser("check", help="Validate readability and list counts; write JSON to --out or print")
    # For check, reuse the common filters but drop chunk and diff-summary flags as they are irrelevant
    pc.add_argument("root", nargs="?", default=".", help="Root directory")
    pc.add_argument("--profile", default="minimal", choices=list(PROFILE_INCLUDES.keys()), help="Inclusion profile")
    pc.add_argument("--include", action="append", default=None, help="Additional include regex (repeatable)")
    pc.add_argument("--exclude", action="append", default=None, help="Additional exclude regex (repeatable)")
    pc.add_argument("--out", default=None, help="Output JSON path (optional)")
    pc.set_defaults(func=run_check)

    return p

if __name__ == "__main__":
    parser = build_arg_parser()
    # Pre-parse just for --version
    if "--version" in sys.argv:
        print(f"paleae {paleae_VERSION}")
        sys.exit(0)
    # Parse known args first to check for --estimate (legacy hook; currently no TokenEstimator impl)
    args, remaining_argv = parser.parse_known_args()
    if getattr(args, 'estimate', False):
        print("[paleae] Note: --estimate hook present but TokenEstimator is not implemented in this build.", file=sys.stderr)
        sys.exit(0)
    # Parse full args (including subcommand)
    args = parser.parse_args(remaining_argv)
    if hasattr(args, 'func'):
        rc = args.func(args)
        sys.exit(rc)
    else:
        parser.print_help()
        sys.exit(1)
